import os
import re
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename

extractos_bp = Blueprint('extractos_bp', __name__)

# Carpetas de trabajo
UPLOAD_FOLDER = '/tmp/uploads'
OUTPUT_FOLDER = '/tmp/output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def process_excel_file(input_filepath, original_filename):
    """
    Procesa el archivo Excel según la lógica proporcionada y devuelve el path
    del archivo modificado.
    """
    # Rutas de los archivos
    archivo_excel_Banco = input_filepath
    output_filename = f"modificado_{secure_filename(original_filename)}"
    archivo_excel_Banco_copia = os.path.join(OUTPUT_FOLDER, output_filename)
    
    # Leer el archivo Excel copiado con pandas
    try:
        # Copiar el archivo original a un archivo de copia
        shutil.copy(archivo_excel_Banco, archivo_excel_Banco_copia)
        # Intentamos cargar la hoja "Movimientos Históricos"
        df_banco = pd.read_excel(archivo_excel_Banco_copia, sheet_name="Movimientos Históricos", header=6)
    except ValueError:
        # Si no existe la hoja, mostramos un mensaje de error y retornamos None
        print(f"La hoja 'Movimientos Históricos' no se encuentra en el archivo: {original_filename}")
        return None

    # Limpiar los nombres de las columnas (eliminar espacios)
    df_banco.columns = df_banco.columns.str.strip()

    # Expresiones regulares para las palabras clave
    patrones_incluir = [r'Iva', r'Ley', r'Perc', r'Misiones', r'Sirc', r'OG-G']
    patron_com = r'Com'
    excluir_com = [r'REC 0']
    patrones_og = [r'OG-T', r'OG - T', r'OG-D', r'OG - D']

    # Función para verificar si "Com" está en el Concepto, y si no tiene "OG" o "REC"
    def agregar_aj(concepto):
        if any(re.search(p, concepto, re.IGNORECASE) for p in patrones_incluir):
            if any(re.search(og, concepto, re.IGNORECASE) for og in patrones_og):
                return ''  # Eliminar "AJ"
            return 'AJ'
        if re.search(patron_com, concepto, re.IGNORECASE):
            if any(re.search(excluir, concepto, re.IGNORECASE) for excluir in excluir_com):
                return ''
            if any(re.search(og, concepto, re.IGNORECASE) for og in patrones_og):
                return ''
            return 'AJ'
        return ''

    # Aplicamos la función a la columna "Concepto" y creamos la nueva columna "AJ"
    df_banco['AJ'] = df_banco['Concepto'].apply(lambda x: agregar_aj(str(x)))

    # Guardamos el archivo modificado con la nueva columna "AJ" en el archivo copiado
    df_banco.to_excel(archivo_excel_Banco_copia, index=False)

    # Crear la hoja de salida (copia del archivo Excel)
    wb = load_workbook(archivo_excel_Banco_copia)

    # Crear una nueva hoja de 'Extracto'
    sheet_extracto = wb.create_sheet('Extracto')

    # Información estática para las primeras filas
    info_extracto = [
        ('324200010', 'Gastos bancarios'),
        ('110401001', 'Iva'),
        ('324200010', 'Gastos bancarios'),
        ('110401010', 'iva 10,5'),
        ('110401014', 'Percepciones iva'),
        ('110401201', 'Percepciones iibb'),
        ('110401201', 'Percepciones misiones'),
        ('110401201', 'Sircreb'),
        ('110401011', 'Impuesto al cheque'),
        ('326001003', 'Impuesto del y crep'),
        ('-', 'Total')
    ]

    # Escribir la información estática en las primeras filas
    for i, (codigo, descripcion) in enumerate(info_extracto, start=1):
        sheet_extracto.cell(row=i, column=1, value=codigo)
        sheet_extracto.cell(row=i, column=2, value=descripcion)

    # Función para sumar valores con condiciones específicas (en columnas 'Débito' o 'Crédito')
    def sumar_conceptos(df, columna, concepto_regex):
        if columna == 'Débito':
            column_values = df['Débito']
        elif columna == 'Crédito':
            column_values = df['Crédito']
        else:
            raise ValueError("La columna debe ser 'Débito' o 'Crédito'.")
        mask = df['Concepto'].str.contains(concepto_regex, flags=re.IGNORECASE, na=False)
        filtered_values = column_values[mask]
        return filtered_values.sum()

    # Ajustar el ancho de las columnas en Sheet1
    sheet1 = wb['Sheet1']
    sheet1.column_dimensions['A'].width = 12
    sheet1.column_dimensions['B'].width = 12
    sheet1.column_dimensions['C'].width = 31
    sheet1.column_dimensions['D'].width = 11
    sheet1.column_dimensions['E'].width = 20
    sheet1.column_dimensions['F'].width = 39
    sheet1.column_dimensions['G'].width = 13
    sheet1.column_dimensions['H'].width = 13
    sheet1.column_dimensions['I'].width = 72
    sheet1.column_dimensions['J'].width = 33
    sheet1.column_dimensions['K'].width = 9

    sheet1.auto_filter.ref = sheet1.dimensions

    # Ajustar el ancho de las columnas en Extracto
    Extracto_sheet = wb['Extracto']
    Extracto_sheet.column_dimensions['A'].width = 15
    Extracto_sheet.column_dimensions['B'].width = 20
    Extracto_sheet.column_dimensions['C'].width = 15
    Extracto_sheet.auto_filter.ref = Extracto_sheet.dimensions

    # Crear un estilo para formato de contabilidad
    contabilidad_style = NamedStyle(name="contabilidad_style", number_format="\"$\"#,##0.00")
    for row in sheet_extracto.iter_rows(min_row=1, min_col=3, max_col=3, max_row=13):
        for cell in row:
            cell.style = contabilidad_style

    # Calcular las filas con los conceptos solo para las filas que tienen 'AJ'
    iva_tasa_gra = (sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'Iva tasa gra') +
                     sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'Iva tasa gra'))
    sheet_extracto.cell(row=2, column=3, value=iva_tasa_gra * -1)

    sheet_extracto.cell(row=1, column=3, value=(iva_tasa_gra / 0.21) * -1)

    iva_10_5 = (sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'Iva tasa red') +
                sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'Iva tasa red'))
    sheet_extracto.cell(row=4, column=3, value=iva_10_5 * -1)

    sheet_extracto.cell(row=3, column=3, value=(iva_10_5 / 0.105) * -1)

    percepciones_iva = (sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'Percep. Iva') +
                        sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'Percep. Iva') +
                        sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'Percepcion I') +
                        sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'Percepcion I') +
                        sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'Recaud') +
                        sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'Recaud'))
    sheet_extracto.cell(row=5, column=3, value=percepciones_iva * -1)

    percepciones_iibb = (sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'Perc.Caba') +
                         sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'Perc.Caba'))
    sheet_extracto.cell(row=6, column=3, value=percepciones_iibb * -1)

    percepciones_misiones = (sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'misiones') +
                              sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'misiones'))
    sheet_extracto.cell(row=7, column=3, value=percepciones_misiones * -1)

    sircreb = (sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', 'Sirc') +
               sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', 'Sirc'))
    sheet_extracto.cell(row=8, column=3, value=sircreb * -1)

    formula_sum = "=C1+C2+C3+C4+C5+C6+C7+C8+C9+C10"
    sheet_extracto.cell(row=11, column=3, value=formula_sum)

    sumatoria_aj = (sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Crédito', '') +
                    sumar_conceptos(df_banco[df_banco['AJ'] == 'AJ'], 'Débito', ''))
    sheet_extracto.cell(row=12, column=3, value=sumatoria_aj * -1)

    formula_dif = "=C11-C12"
    sheet_extracto.cell(row=13, column=3, value=formula_dif)

    wb.save(archivo_excel_Banco_copia)
    # os.startfile(archivo_excel_Banco_copia)  # No se usa en web
    print(f"Archivo modificado y guardado como: {archivo_excel_Banco_copia}")
    return archivo_excel_Banco_copia

# Rutas de Flask para Extractos
@extractos_bp.route('/extracto', methods=['POST'])
def extracto():
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    resultado = process_excel_file(filepath, file.filename)
    if resultado is None:
        return jsonify({'error': 'Error al procesar el archivo'}), 500

    return jsonify({'mensaje': 'Archivo procesado correctamente', 'archivo': resultado})
