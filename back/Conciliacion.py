import os
import shutil
import pandas as pd
import re
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, redirect, url_for
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

conciliacion_bp = Blueprint('conciliacion_bp', __name__)

# Crear las carpetas si no existen.
UPLOAD_FOLDER = '/tmp/uploads'
OUTPUT_FOLDER = '/tmp/output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def procesar_archivos(archivo_original, archivo_extracto):
    archivo_copia = os.path.join(OUTPUT_FOLDER, "ConciliacionPrueba.xlsx")
    shutil.copy(archivo_original, archivo_copia)

    try:
        wb = load_workbook(archivo_copia)
        ws = wb.active  # Tomamos la primera hoja

    except Exception as e:
        return f"Error al abrir el archivo: {str(e)}"

    ws["A3"] = "Origen"
    ws["I3"] = "Minuta"
    ws["P3"] = "Valor Neto"
    ws["Q3"] = "Cruce"

    ws.auto_filter.ref = f"A3:{ws.cell(row=3, column=ws.max_column+1).coordinate}"

    for row in range(4, ws.max_row + 1):
        ws[f"A{row}"] = "CONTABILIDAD"

    ws.insert_cols(11)
    ws["K3"] = "Mes"

    def obtener_valor_si_none(valor):
        return valor if valor is not None else 0

    def convertir_a_fecha(valor):
        try:
            return datetime.strptime(valor, "%d-%m-%Y")
        except (ValueError, TypeError):
            return None

    try:
        wb_extracto = load_workbook(archivo_extracto)
        ws_extracto = wb_extracto["Sheet1"]
    except Exception as e:
        return f"Error al abrir extracto: {str(e)}"

    for row in range(2, ws_extracto.max_row + 1):
        fecha = convertir_a_fecha(ws_extracto[f"A{row}"].value)
        tipo = ws_extracto[f"K{row}"].value
        concepto = ws_extracto[f"C{row}"].value
        credito = ws_extracto[f"G{row}"].value
        debito = ws_extracto[f"H{row}"].value

        new_row = ws.max_row + 1

        ws[f"A{new_row}"] = "EXTRACTO"
        ws[f"J{new_row}"] = fecha
        ws[f"L{new_row}"] = tipo
        ws[f"M{new_row}"] = concepto
        ws[f"O{new_row}"] = obtener_valor_si_none(credito)
        ws[f"P{new_row}"] = obtener_valor_si_none(debito) * -1

    for row in range(4, ws.max_row + 1):
        fecha = ws[f"J{row}"].value
        if isinstance(fecha, datetime):
            ws[f"K{row}"] = fecha.month
            ws[f"J{row}"].number_format = "DD/MM/YYYY"
        else:
            ws[f"K{row}"] = ""

    for row in range(4, ws.max_row + 1):
        origen = ws[f"A{row}"].value
        operacion = ws[f"M{row}"].value
        debe = ws[f"O{row}"].value
        tipo = ws[f"L{row}"].value

        if origen == "EXTRACTO":
            if re.search(r'\.CH', operacion):
                ws[f"L{row}"] = "DE"
            elif tipo is None or tipo == "":
                if debe > 0:
                    ws[f"L{row}"] = "RC"
                else:
                    ws[f"L{row}"] = "OP"

    for row in range(4, ws.max_row + 1):
        valor_neto = (ws[f"O{row}"].value or 0) - (ws[f"P{row}"].value or 0)
        ws[f"Q{row}"].value = valor_neto
        ws[f"R{row}"].value = valor_neto if ws[f"A{row}"].value == "CONTABILIDAD" else -valor_neto

    # Guardar el orden original antes de ordenar
    orden_original = [(i, ws[f"A{i}"].value, ws[f"J{i}"].value) for i in range(4, ws.max_row + 1)]

    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[2]
    df = df.iloc[3:]
    df = df.sort_values(by=["Tipo", "Valor Neto"], ascending=[False, True])

    for i, row in enumerate(df.itertuples(index=False), start=4):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    green_cells = set()

    for row in range(4, ws.max_row):
        tipo_actual = ws[f"L{row}"].value
        cruce_actual = ws[f"R{row}"].value
        tipo_siguiente = ws[f"L{row+1}"].value
        cruce_siguiente = ws[f"R{row+1}"].value

        if tipo_actual == tipo_siguiente and (cruce_actual + cruce_siguiente) == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill
                ws.cell(row=row+1, column=col).fill = green_fill
                green_cells.add(row)
                green_cells.add(row+1)

    # Restaurar el orden original después de la conciliación
    for idx, origen, fecha in orden_original:
        ws[f"A{idx}"] = origen
        ws[f"J{idx}"] = fecha

    # CREACIÓN DE LA HOJA "TRANSITORIAS"
    ws_transitorias = wb.create_sheet("Transitorias")
    
    ws_transitorias["A1"] = "Hoja generada con Transitorias"

    for col in range(1, ws.max_column + 1):
        ws_transitorias.cell(row=3, column=col, value=ws.cell(row=3, column=col).value)

    transitorias_row = 4

    for row in range(4, ws.max_row + 1):
        if row not in green_cells:
            for col in range(1, ws.max_column + 1):
                ws_transitorias.cell(row=transitorias_row, column=col, value=ws.cell(row=row, column=col).value)
            transitorias_row += 1

    for row in range(4, transitorias_row):
        for col in ["O", "P", "Q", "R"]:
            ws_transitorias[f"{col}{row}"].number_format = '_-$ * #,##0.00_-;-$ * #,##0.00_-;_-$ * "-"??_-;_-@_-'
        ws_transitorias[f"J{row}"].number_format = 'DD/MM/YYYY'

    column_widths = {"A": 25, "B": 12, "C": 17, "D": 12, "E": 21, "F": 12, "G": 34, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12, "M": 16, "N": 31, "O": 18, "P": 18, "Q": 18, "R": 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
        ws_transitorias.column_dimensions[col].width = width

    ws.sheet_view.rightToLeft = False
    ws.sheet_view.topLeftCell = "A1"
    ws.freeze_panes = "D4"

    ws_transitorias.sheet_view.rightToLeft = False
    ws_transitorias.sheet_view.topLeftCell = "A1"
    ws_transitorias.freeze_panes = "D4"

    wb.save(archivo_copia)
    return archivo_copia

@conciliacion_bp.route('/conciliacion', methods=['GET', 'POST'])
def conciliacion():
    if request.method == 'POST':
        # Verificar que se hayan enviado ambos archivos
        if 'archivo_original' not in request.files or 'archivo_extracto' not in request.files:
            return jsonify({'error': 'No se envió uno o ambos archivos'}), 400

        archivo_original = request.files['archivo_original']
        archivo_extracto = request.files['archivo_extracto']

        if archivo_original.filename == '' or archivo_extracto.filename == '':
            return jsonify({'error': 'Nombre de archivo vacío en uno de los archivos'}), 400

        # Guardar archivos
        filepath_original = os.path.join(UPLOAD_FOLDER, archivo_original.filename)
        filepath_extracto = os.path.join(UPLOAD_FOLDER, archivo_extracto.filename)
        archivo_original.save(filepath_original)
        archivo_extracto.save(filepath_extracto)

        # Aquí puedes llamar a tu función de procesamiento que use ambos archivos
        # Procesar los archivos
        resultado = procesar_archivos(filepath_original, filepath_extracto)
        if isinstance(resultado, str) and resultado.startswith("Error"):
            return jsonify({'error': resultado}), 500

        # Redirigir a la ruta intermedia que disparará la descarga y luego redirigirá al menú
        return redirect(url_for('resultado', filename=os.path.basename(resultado)))
    
    # Si es GET, renderiza el formulario
    return render_template('conciliacion.html')
